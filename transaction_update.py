# Libraries
import pandas as pd
import os
import openpyxl


# Path where the excel sheets exist.
path = "C:/Devish/transaction history"
# Returns a list of excel sheets present in the path.
files = os.listdir(path)


def reading_excel(file_index):
    # Reading the excel file besed on the file_index.
    dataframe = pd.read_excel(path + '/' + files[file_index], engine= "openpyxl")
    # print(dataframe)
    return(dataframe)


def selecting_right_file(year):
    # Opens the file based on the current year, if the year does not exist it will come out of the loop.
    # year = "2023"
    for i, file in enumerate(files):
        if(year == file[20:24]):
            return(i)
    # If the excel sheet of the current year is not present it will create a new excel sheet.
    intial_dataframe = pd.DataFrame({"Date" : [], "Time" : [], "Amount of transaction" : [], "Initial amount" : [], "Final Amount" : [], "Mode of transaction" : [], "Tansaction id" : [],"Alert" : [], "Error: Reason" : []})
    new_excel = pd.ExcelWriter(path + "/" + f"transaction_history({year}).xlsx", engine = 'xlsxwriter')
    intial_dataframe.to_excel(new_excel, sheet_name = "Sheet1", index = False)
    new_excel.save()
    new_excel.close()
    files.append(f"transaction_history({year}).xlsx")
    return(len(files) - 1)


def appending_excel(file_index, append_values):
    dataframe = pd.read_excel(path + '/' + files[file_index], engine= "openpyxl")
    input_dataframe = pd.DataFrame(append_values)
    dataframe = pd.concat([dataframe, input_dataframe], ignore_index = True)
    dataframe.to_excel(path + '/' + files[file_index], sheet_name='Sheet1', index = False)
    

# This metheod adds color to column Alerts.
def alert_color_fix(dataframe, file_index):
    excel = openpyxl.load_workbook(path + '/' + files[file_index])
    read_excel_sheet = excel['Sheet1']
    errors = dataframe.iloc[:, [-1]].values
    for i, j in enumerate(errors):
        cell_color = openpyxl.styles.PatternFill(patternType='solid', fgColor='35FC03')
        if(j != "NONE"):
            # For red color if cell alert is yes in the transaction.
            cell_color = openpyxl.styles.PatternFill(patternType='solid', fgColor='FC2C03')
        read_excel_sheet["H" + str(i + 2)].fill = cell_color
    excel.save(path + '/' + files[file_index])


# This mtheod is to check if pervious year excel is present or not.
def pervious_year(year):
    for file in files:
        if(year == int(file[20:24])):
            return(True)
    return(False)